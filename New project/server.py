import hashlib
import base64
import hmac
import json
import mimetypes
import secrets
import sqlite3
import traceback
import sys
import struct
import threading
from cgi import FieldStorage
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

import openpyxl

from auth import (
    AuthHandler,
    ROLE_ALLIANCEADMIN,
    ROLE_GUEST,
    ROLE_SUPERADMIN,
    ROLE_VERIFIEDUSER,
    get_current_auth,
    initialize_auth_database,
    update_user_sessions_for_user,
)


BASE_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = BASE_DIR / "public"
DATA_DIR = BASE_DIR / "data"
UPLOADS_DIR = PUBLIC_DIR / "uploads" / "member-screenshots"
DB_PATH = DATA_DIR / "alliance.db"
HOST = "127.0.0.1"
PORT = 8000
SESSION_COOKIE = "alliance_session"
SESSION_TTL_SECONDS = 60 * 60 * 12

sessions = {}

# WebSocket clients connected for real-time melon updates
ws_clients = set()
ws_clients_lock = threading.Lock()
WS_MAGIC_GUID = "258EAFA5-E914-47DA-95CA-C5AB0DC85B11"

# Thread pool for async database writes
db_executor = ThreadPoolExecutor(max_workers=2)


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_scaled_number(value, field_name="数值"):
    raw = str(value or "").strip()
    if not raw:
        return 0

    normalized = raw.replace(",", "").replace("％", "%").replace(" ", "")
    if normalized.endswith("%"):
        normalized = normalized[:-1]

    composite_multiplier = None
    for composite_suffix, multiplier_value in (("万亿", 1_000_000_000_000), ("萬億", 1_000_000_000_000)):
        if normalized.endswith(composite_suffix):
            composite_multiplier = multiplier_value
            normalized = normalized[:-len(composite_suffix)]
            break

    units = {
        "k": 1_000,
        "K": 1_000,
        "千": 1_000,
        "w": 10_000,
        "W": 10_000,
        "万": 10_000,
        "亿": 100_000_000,
    }

    multiplier = composite_multiplier or 1
    if composite_multiplier is None:
        suffix = normalized[-1:] if normalized else ""
        if suffix in units:
            multiplier = units[suffix]
            normalized = normalized[:-1]

    try:
        return round(float(normalized or "0") * multiplier, 4)
    except ValueError as exc:
        raise ValueError(f"{field_name}格式不正确，支持小数，也支持万、亿、万亿等单位") from exc


def hash_password(password: str, salt: str) -> str:
    hashed = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return hashed.hex()


def open_db():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def build_guild_display_name(guild_code, guild_prefix, guild_name):
    parts = [str(guild_code or "").strip(), str(guild_prefix or "").strip(), str(guild_name or "").strip()]
    return " ".join(part for part in parts if part)


def build_guild_key(guild_code, guild_prefix, guild_name):
    return "|".join([
        str(guild_code or "").strip(),
        str(guild_prefix or "").strip(),
        str(guild_name or "").strip(),
    ])


def upsert_guild_registry(connection, member, leader_name=None):
    guild_key = build_guild_key(member.get("guild_code", ""), member.get("guild_prefix", ""), member.get("guild", ""))
    if not guild_key:
        return
    connection.execute(
        """
        INSERT INTO guild_registry (
            guild_key, alliance, hill, guild_code, guild_prefix, guild, guild_power, leader_name, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(guild_key) DO UPDATE SET
            alliance = excluded.alliance,
            hill = excluded.hill,
            guild_code = excluded.guild_code,
            guild_prefix = excluded.guild_prefix,
            guild = excluded.guild,
            guild_power = excluded.guild_power,
            leader_name = excluded.leader_name,
            updated_at = excluded.updated_at
        """,
        (
            guild_key,
            member.get("alliance", "🔮联盟"),
            member.get("hill", "默认山头"),
            member.get("guild_code", ""),
            member.get("guild_prefix", ""),
            member.get("guild", ""),
            int(member.get("guild_power", 0) or 0),
            leader_name if leader_name is not None else member.get("name", ""),
            member.get("updated_at") or member.get("created_at") or now_text(),
        ),
    )


def guild_exists(connection, guild_code, guild_prefix, guild_name, exclude_key=None):
    conditions = []
    params = []
    if str(guild_code or "").strip():
        conditions.append("guild_code = ?")
        params.append(str(guild_code).strip())
    else:
        conditions.append("guild_prefix = ? AND guild = ?")
        params.extend([str(guild_prefix or "").strip(), str(guild_name or "").strip()])
    if exclude_key:
        conditions.append("guild_key <> ?")
        params.append(exclude_key)
    query = f"SELECT guild_key FROM guild_registry WHERE {' AND '.join(conditions)} LIMIT 1"
    return connection.execute(query, tuple(params)).fetchone()


def member_name_exists(connection, guild_code, guild_prefix, guild_name, member_name, exclude_id=None):
    query = """
        SELECT id FROM members
        WHERE guild_code = ? AND guild_prefix = ? AND guild = ? AND name = ?
    """
    params = [str(guild_code or "").strip(), str(guild_prefix or "").strip(), str(guild_name or "").strip(), str(member_name or "").strip()]
    if exclude_id is not None:
        query += " AND id <> ?"
        params.append(exclude_id)
    query += " LIMIT 1"
    return connection.execute(query, tuple(params)).fetchone()


def initialize_database():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    with open_db() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                salt TEXT NOT NULL,
                display_name TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS announcements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                category TEXT NOT NULL,
                created_at TEXT NOT NULL,
                author TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS guild_registry (
                guild_key TEXT PRIMARY KEY,
                alliance TEXT NOT NULL,
                hill TEXT NOT NULL DEFAULT '默认山头',
                guild_code TEXT NOT NULL DEFAULT '',
                guild_prefix TEXT NOT NULL DEFAULT '',
                guild TEXT NOT NULL,
                guild_power INTEGER NOT NULL DEFAULT 0,
                leader_name TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                alliance TEXT NOT NULL,
                hill TEXT NOT NULL DEFAULT '默认山头',
                guild_code TEXT NOT NULL DEFAULT '',
                guild_prefix TEXT NOT NULL DEFAULT '',
                guild_power INTEGER NOT NULL DEFAULT 0,
                guild TEXT NOT NULL,
                name TEXT NOT NULL,
                role TEXT NOT NULL,
                realm TEXT NOT NULL,
                power INTEGER NOT NULL DEFAULT 0,
                hp INTEGER NOT NULL DEFAULT 0,
                attack INTEGER NOT NULL DEFAULT 0,
                defense INTEGER NOT NULL DEFAULT 0,
                speed INTEGER NOT NULL DEFAULT 0,
                bonus_damage INTEGER NOT NULL DEFAULT 0,
                damage_reduction INTEGER NOT NULL DEFAULT 0,
                pet TEXT NOT NULL,
                note TEXT NOT NULL DEFAULT '',
                screenshot_path TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            """
        )

        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS member_cert_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                member_id INTEGER NOT NULL,
                alliance TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL,
                reviewed_at TEXT,
                reviewer_id INTEGER
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS admin_role_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                alliance TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                is_read INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                read_at TEXT,
                reviewed_at TEXT,
                reviewer_id INTEGER
            )
            """
        )
        admin_role_request_columns = [row["name"] for row in connection.execute("PRAGMA table_info(admin_role_requests)").fetchall()]
        if "is_read" not in admin_role_request_columns:
            connection.execute("ALTER TABLE admin_role_requests ADD COLUMN is_read INTEGER NOT NULL DEFAULT 0")
        if "read_at" not in admin_role_request_columns:
            connection.execute("ALTER TABLE admin_role_requests ADD COLUMN read_at TEXT")

        admin_count = connection.execute("SELECT COUNT(*) AS count FROM admins").fetchone()["count"]
        if admin_count == 0:
            salt = secrets.token_hex(16)
            connection.execute(
                """
                INSERT INTO admins (username, password_hash, salt, display_name, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                ("admin", hash_password("123456", salt), salt, "联盟管理员", now_text()),
            )

        columns = [row["name"] for row in connection.execute("PRAGMA table_info(members)").fetchall()]
        if "hill" not in columns:
            connection.execute("ALTER TABLE members ADD COLUMN hill TEXT NOT NULL DEFAULT '默认山头'")
        if "guild_code" not in columns:
            connection.execute("ALTER TABLE members ADD COLUMN guild_code TEXT NOT NULL DEFAULT ''")
        if "guild_prefix" not in columns:
            connection.execute("ALTER TABLE members ADD COLUMN guild_prefix TEXT NOT NULL DEFAULT ''")
        if "guild_power" not in columns:
            connection.execute("ALTER TABLE members ADD COLUMN guild_power INTEGER NOT NULL DEFAULT 0")
        if "bonus_damage" not in columns:
            connection.execute("ALTER TABLE members ADD COLUMN bonus_damage INTEGER NOT NULL DEFAULT 0")
        if "damage_reduction" not in columns:
            connection.execute("ALTER TABLE members ADD COLUMN damage_reduction INTEGER NOT NULL DEFAULT 0")
        if "verified" not in columns:
            connection.execute("ALTER TABLE members ADD COLUMN verified INTEGER NOT NULL DEFAULT 0")
        if "screenshot_path" not in columns:
            connection.execute("ALTER TABLE members ADD COLUMN screenshot_path TEXT NOT NULL DEFAULT ''")

        announcement_columns = [row["name"] for row in connection.execute("PRAGMA table_info(announcements)").fetchall()]
        if "author" not in announcement_columns:
            connection.execute("ALTER TABLE announcements ADD COLUMN author TEXT NOT NULL DEFAULT ''")

        connection.execute("CREATE INDEX IF NOT EXISTS idx_guild_registry_code_lookup ON guild_registry(guild_code)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_members_guild_name_lookup ON members(guild_code, guild_prefix, guild, name)")

        announcement_count = connection.execute("SELECT COUNT(*) AS count FROM announcements").fetchone()["count"]
        if announcement_count == 0:
            seed_announcements = [
                ("联盟招新开启", "本周开放 2 个妖盟名额，要求活跃、配合填表。", "公告"),
                ("周日晚联盟战", "请各妖盟于 20:00 前同步主力战力和灵兽配置。", "公告"),
                ("天狐妖盟昨晚冲榜成功", "前三主力全部破 30 万战，今天已经坐稳联盟第一梯队。", "瓜棚"),
                ("玄龟妖盟准备换阵", "据说正在测试双反击流，有望下周冲进前二。", "瓜棚"),
            ]
            connection.executemany(
                "INSERT INTO announcements (title, content, category, created_at, author) VALUES (?, ?, ?, ?, ?)",
                [(title, content, category, now_text()) for title, content, category in seed_announcements],
            )

        member_count = connection.execute("SELECT COUNT(*) AS count FROM members").fetchone()["count"]
        if member_count == 0:
            timestamp = now_text()
            seed_members = [
                ("🔮联盟", "蓬莱1-方丈12", "833", "长虹山", 865200, "天狐妖盟", "青玄子", "盟主", "化神后期", 328000, 98000, 23500, 16800, 9200, "应龙", "主修暴击流，负责联盟指挥。", timestamp, timestamp),
                ("🔮联盟", "方丈13-瀛洲4", "612", "望月山", 612800, "玄龟妖盟", "白芷", "副盟主", "元婴圆满", 285600, 86200, 21000, 15200, 10100, "九尾狐", "负责资源统计与招新审核。", timestamp, timestamp),
                ("🔮联盟", "蓬莱1-方丈12", "501", "赤羽山", 501300, "赤羽妖盟", "曜尘", "长老", "元婴后期", 243500, 79800, 18700, 14500, 8800, "玄龟", "主抗伤阵容，联盟战前排。", timestamp, timestamp),
            ]
            connection.executemany(
                """
                INSERT INTO members (
                    alliance, hill, guild_code, guild_prefix, guild_power, guild, name, role, realm, power, hp, attack, defense, speed, pet, note, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                seed_members,
            )

        registry_count = connection.execute("SELECT COUNT(*) AS count FROM guild_registry").fetchone()["count"]
        if registry_count == 0:
            rows = [dict(row) for row in connection.execute("SELECT * FROM members ORDER BY updated_at DESC, id DESC").fetchall()]
            grouped = {}
            for row in rows:
                key = build_guild_key(row.get("guild_code", ""), row.get("guild_prefix", ""), row.get("guild", ""))
                if not key:
                    continue
                entry = grouped.setdefault(key, {**row})
                entry["guild_power"] = max(int(entry.get("guild_power", 0) or 0), int(row.get("guild_power", 0) or 0))
                if row.get("role") == "盟主" and row.get("name"):
                    entry["name"] = row["name"]
            for entry in grouped.values():
                upsert_guild_registry(connection, entry)

        connection.execute("UPDATE members SET alliance = '🔮联盟' WHERE alliance = '青云联盟'")

        connection.commit()


def cleanup_sessions():
    current = datetime.now().timestamp()
    expired = [token for token, value in sessions.items() if current - value["created_at"] > SESSION_TTL_SECONDS]
    for token in expired:
        sessions.pop(token, None)


class AllianceHandler(BaseHTTPRequestHandler):
    server_version = "AlliancePortal/1.0"

    def do_GET(self):
        self.run_safely(self._do_GET_impl)

    def do_POST(self):
        self.run_safely(self._do_POST_impl)

    def do_PUT(self):
        self.run_safely(self._do_PUT_impl)

    def do_DELETE(self):
        self.run_safely(self._do_DELETE_impl)

    def _do_GET_impl(self):
        cleanup_sessions()
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/auth/"):
            AuthHandler(self).handle()
            return
        if parsed.path == "/ws/melon":
            self.handle_melon_websocket_upgrade()
            return
        if parsed.path.startswith("/api/"):
            self.handle_api_get(parsed)
            return
        self.serve_static(parsed.path)

    def _do_POST_impl(self):
        cleanup_sessions()
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/auth/"):
            AuthHandler(self).handle()
            return
        if parsed.path.startswith("/api/"):
            self.handle_api_post(parsed)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def _do_PUT_impl(self):
        cleanup_sessions()
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            self.handle_api_put(parsed)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def _do_DELETE_impl(self):
        cleanup_sessions()
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            self.handle_api_delete(parsed)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def run_safely(self, handler):
        try:
            handler()
        except BrokenPipeError:
            raise
        except Exception as exc:
            sys.stderr.write(f"[server-error] {self.command} {self.path}: {exc}\n")
            traceback.print_exc()
            if getattr(self, "wfile", None) and not self.wfile.closed:
                try:
                    self.send_json({"error": f"服务器内部错误: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                except Exception:
                    pass

    def serve_static(self, path: str):
        target = PUBLIC_DIR / path.lstrip("/")
        if path in ("", "/"):
            target = PUBLIC_DIR / "auth.html"

        if target.is_dir():
            target = target / "index.html"

        try:
            resolved = target.resolve()
            if not str(resolved).startswith(str(PUBLIC_DIR.resolve())) or not resolved.exists():
                raise FileNotFoundError
        except FileNotFoundError:
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        content_type = "text/plain; charset=utf-8"
        if resolved.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        elif resolved.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif resolved.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        elif resolved.suffix == ".json":
            content_type = "application/json; charset=utf-8"
        elif resolved.suffix in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
            content_type = mimetypes.guess_type(str(resolved))[0] or "application/octet-stream"

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.end_headers()
        self.wfile.write(resolved.read_bytes())

    def handle_api_get(self, parsed):
        if parsed.path == "/api/health":
            self.send_json({"status": "ok", "time": now_text()})
            return
        if parsed.path == "/api/me":
            self.send_json(get_current_auth(self))
            return
        if parsed.path == "/api/dashboard":
            self.send_json(self.build_dashboard())
            return
        if parsed.path == "/api/profile/me":
            self.send_json(self.get_current_user_profile())
            return
        if parsed.path == "/api/members":
            query = parse_qs(parsed.query)
            self.send_json(self.list_members(query))
            return
        if parsed.path == "/api/announcements":
            self.send_json(self.list_announcements())
            return
        if parsed.path == "/api/member-cert-requests":
            current = get_current_auth(self)
            if not current.get("authenticated"):
                self.send_json({"error": "请先登录"}, status=HTTPStatus.UNAUTHORIZED)
                return
            query = parse_qs(parsed.query)
            member_id = query.get("member_id", [""])[0].strip() or None
            self.send_json(self.list_member_cert_requests(current, member_id=member_id))
            return
        if parsed.path == "/api/member-cert-requests/mine":
            current = get_current_auth(self)
            if not current.get("authenticated"):
                self.send_json({"error": "请先登录"}, status=HTTPStatus.UNAUTHORIZED)
                return
            self.send_json(self.list_member_cert_requests(current, mine=True))
            return
        if parsed.path == "/api/admin-role-requests":
            user = self.require_permission("manage_roles")
            if not user:
                return
            query = parse_qs(parsed.query)
            mark_read = query.get("mark_read", ["0"])[0] in {"1", "true", "yes"}
            self.send_json(self.list_admin_role_requests(mark_read=mark_read))
            return
        if parsed.path.startswith("/api/members/") and parsed.path.endswith("/screenshot"):
            member_id = parsed.path.strip("/").split("/")[2]
            self.send_json(self.get_member_screenshot(member_id))
            return
        self.send_json({"error": "接口不存在"}, status=HTTPStatus.NOT_FOUND)

    def handle_melon_websocket_upgrade(self):
        upgrade_header = (self.headers.get("Upgrade") or "").lower()
        connection_header = (self.headers.get("Connection") or "").lower()
        key = self.headers.get("Sec-WebSocket-Key")
        version = self.headers.get("Sec-WebSocket-Version")
        if upgrade_header != "websocket" or "upgrade" not in connection_header or not key or version != "13":
            self.send_error(HTTPStatus.BAD_REQUEST, "Invalid WebSocket handshake")
            return

        accept_value = base64.b64encode(
            hashlib.sha1((key + WS_MAGIC_GUID).encode("utf-8")).digest()
        ).decode("ascii")
        self.send_response(HTTPStatus.SWITCHING_PROTOCOLS)
        self.send_header("Upgrade", "websocket")
        self.send_header("Connection", "Upgrade")
        self.send_header("Sec-WebSocket-Accept", accept_value)
        self.end_headers()

        client = MelonWebSocketClient(self.connection, self.client_address)
        register_ws_client(client)
        try:
            client.read_loop()
        finally:
            unregister_ws_client(client)

    def handle_api_post(self, parsed):
        if parsed.path == "/api/guilds":
            payload = self.read_json()
            user = self.require_permission("manage_guilds", payload.get("alliance", ""))
            if not user:
                return
            self.create_guild(payload)
            return
        if parsed.path == "/api/members":
            payload = self.read_json()
            user = self.require_permission("manage_members", payload.get("alliance", ""))
            if not user:
                return
            self.create_member(payload)
            return
        if parsed.path == "/api/profile/me/screenshot":
            user = self.require_permission("upload_own_screenshot", allow_admin_account=False)
            if not user:
                return
            self.upload_current_user_screenshot(user)
            return
        if parsed.path == "/api/login":
            AuthHandler(self).login()
            return
        if parsed.path == "/api/logout":
            AuthHandler(self).logout()
            return
        if parsed.path == "/api/announcements":
            user = self.require_permission("manage_announcements")
            if not user:
                return
            payload = self.read_json()
            self.create_announcement(payload)
            return
        if parsed.path.startswith("/api/members/") and parsed.path.endswith("/screenshot"):
            member_id = parsed.path.strip("/").split("/")[2]
            member = self.get_member_item(member_id)
            if not member:
                self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            user = self.require_permission("manage_members", member.get("alliance", ""))
            if not user:
                return
            self.upload_member_screenshot(member_id)
            return
        if parsed.path == "/api/members/import":
            current = get_current_auth(self)
            if not current.get("authenticated"):
                self.send_json({"error": "请先登录"}, status=HTTPStatus.UNAUTHORIZED)
                return
            user = current.get("user") or {}
            if not self.has_permission(user, "manage_members"):
                self.send_json({"error": "当前账号没有执行该操作的权限"}, status=HTTPStatus.FORBIDDEN)
                return
            self.import_members_from_excel()
            return
        if parsed.path == "/api/member-cert-requests":
            current = get_current_auth(self)
            if not current.get("authenticated") or current.get("user", {}).get("role") != ROLE_GUEST:
                self.send_json({"error": "仅 Guest 可发起认证申请"}, status=HTTPStatus.FORBIDDEN)
                return
            payload = self.read_json()
            self.create_member_cert_request(current, payload)
            return
        if parsed.path == "/api/admin-role-requests":
            current = get_current_auth(self)
            if not current.get("authenticated") or current.get("is_admin"):
                self.send_json({"error": "仅普通用户可申请联盟管理员"}, status=HTTPStatus.FORBIDDEN)
                return
            payload = self.read_json()
            self.create_admin_role_request(current, payload)
            return
        if parsed.path.startswith("/api/member-cert-requests/"):
            current = get_current_auth(self)
            if not current.get("authenticated") or not self.has_permission(current.get("user") or {}, "manage_members"):
                self.send_json({"error": "当前账号没有审核认证申请的权限"}, status=HTTPStatus.FORBIDDEN)
                return
            payload = self.read_json()
            request_id = parsed.path.rsplit("/", 1)[-1]
            self.review_member_cert_request(request_id, payload, current)
            return
        if parsed.path.startswith("/api/admin-role-requests/"):
            user = self.require_permission("manage_roles")
            if not user:
                return
            payload = self.read_json()
            request_id = parsed.path.rsplit("/", 1)[-1]
            self.review_admin_role_request(request_id, payload, {"authenticated": True, "user": user})
            return
        if parsed.path == "/api/melon":
            user = self.get_current_user_or_admin()
            if not user:
                self.send_json({"error": "请先认证"}, status=HTTPStatus.UNAUTHORIZED)
                return
            if not self.has_permission(user, "create_posts"):
                self.send_json({"error": "当前账号没有发布瓜棚的权限"}, status=HTTPStatus.FORBIDDEN)
                return
            payload = self.read_json()
            self.create_melon_post(user, payload)
            return
        self.send_json({"error": "接口不存在"}, status=HTTPStatus.NOT_FOUND)

    def handle_api_put(self, parsed):
        if parsed.path.startswith("/api/guilds/"):
            guild_key = unquote(parsed.path.rsplit("/", 1)[-1])
            existing = self.get_guild_registry_item(guild_key)
            if not existing:
                self.send_json({"error": "妖盟不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            payload = self.read_json()
            user = self.require_permission("manage_guilds", existing.get("alliance", ""))
            if not user:
                return
            target_alliance = str(payload.get("alliance", existing.get("alliance", ""))).strip()
            if target_alliance and target_alliance != str(existing.get("alliance", "")).strip() and not self.can_access_alliance(user, target_alliance, "manage_guilds"):
                self.send_json({"error": "当前账号不能把妖盟调整到其他联盟"}, status=HTTPStatus.FORBIDDEN)
                return
            self.update_guild(guild_key, payload)
            return
        if parsed.path == "/api/profile/me":
            user = self.require_permission("edit_own_profile", allow_admin_account=False)
            if not user:
                return
            payload = self.read_json()
            self.update_current_user_profile(user, payload)
            return
        if parsed.path.startswith("/api/members/"):
            member_id = parsed.path.rsplit("/", 1)[-1]
            existing = self.get_member_item(member_id)
            if not existing:
                self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            payload = self.read_json()
            user = self.require_permission("manage_members", existing.get("alliance", ""))
            if not user:
                return
            target_alliance = str(payload.get("alliance", existing.get("alliance", ""))).strip()
            if target_alliance and target_alliance != str(existing.get("alliance", "")).strip() and not self.can_access_alliance(user, target_alliance, "manage_members"):
                self.send_json({"error": "当前账号不能把成员调整到其他联盟"}, status=HTTPStatus.FORBIDDEN)
                return
            self.update_member(member_id, payload)
            return
        if parsed.path.startswith("/api/announcements/"):
            user = self.require_permission("manage_announcements")
            if not user:
                return
            payload = self.read_json()
            announcement_id = parsed.path.rsplit("/", 1)[-1]
            self.update_announcement(announcement_id, payload)
            return
        self.send_json({"error": "接口不存在"}, status=HTTPStatus.NOT_FOUND)

    def handle_api_delete(self, parsed):
        if parsed.path.startswith("/api/guilds/"):
            guild_key = unquote(parsed.path.rsplit("/", 1)[-1])
            existing = self.get_guild_registry_item(guild_key)
            if not existing:
                self.send_json({"error": "妖盟不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            user = self.require_permission("manage_guilds", existing.get("alliance", ""))
            if not user:
                return
            self.delete_guild(guild_key)
            return
        if parsed.path == "/api/profile/me/screenshot":
            user = self.require_permission("upload_own_screenshot", allow_admin_account=False)
            if not user:
                return
            self.delete_current_user_screenshot(user)
            return
        if parsed.path.startswith("/api/members/") and parsed.path.endswith("/screenshot"):
            member_id = parsed.path.strip("/").split("/")[2]
            member = self.get_member_item(member_id)
            if not member:
                self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            user = self.require_permission("manage_members", member.get("alliance", ""))
            if not user:
                return
            self.delete_member_screenshot(member_id)
            return
        if parsed.path.startswith("/api/members/"):
            member_id = parsed.path.rsplit("/", 1)[-1]
            member = self.get_member_item(member_id)
            if not member:
                self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            user = self.require_permission("manage_members", member.get("alliance", ""))
            if not user:
                return
            self.delete_by_id("members", member_id)
            return
        if parsed.path.startswith("/api/announcements/"):
            user = self.require_permission("manage_announcements")
            if not user:
                return
            announcement_id = parsed.path.rsplit("/", 1)[-1]
            self.delete_by_id("announcements", announcement_id)
            return
        if parsed.path.startswith("/api/melon/"):
            user = self.get_current_user_or_admin()
            if not user:
                self.send_json({"error": "请先登录"}, status=HTTPStatus.UNAUTHORIZED)
                return
            melon_id = parsed.path.rsplit("/", 1)[-1]
            self.delete_melon_post(user, melon_id)
            return
        self.send_json({"error": "接口不存在"}, status=HTTPStatus.NOT_FOUND)

    def build_dashboard(self):
        with open_db() as connection:
            members = [dict(row) for row in connection.execute("SELECT * FROM members").fetchall()]
            guild_registry = [dict(row) for row in connection.execute("SELECT * FROM guild_registry").fetchall()]
            announcements = [dict(row) for row in connection.execute("SELECT * FROM announcements ORDER BY id DESC").fetchall()]

        total_power = sum(member["power"] for member in members)
        hill_map = {}
        for registry in guild_registry:
            hill_name = registry["hill"] or "默认山头"
            guild_key = registry["guild_key"]
            hill_entry = hill_map.setdefault(hill_name, {"count": 0, "power": 0, "guilds": {}})
            hill_entry["guilds"].setdefault(guild_key, {
                "name": registry["guild"],
                "code": registry.get("guild_code", ""),
                "prefix": registry.get("guild_prefix", ""),
                "display_name": build_guild_display_name(registry.get("guild_code", ""), registry.get("guild_prefix", ""), registry["guild"]),
                "count": 0,
                "power": 0,
                "custom_power": int(registry.get("guild_power", 0) or 0),
                "top_member": None,
                "leader_name": registry.get("leader_name", ""),
                "updated_at": registry.get("updated_at", ""),
            })

        for member in members:
            hill_name = member["hill"] or "默认山头"
            guild_name = member["guild"]
            guild_code = member.get("guild_code", "")
            guild_prefix = member.get("guild_prefix", "")
            guild_key = build_guild_key(guild_code, guild_prefix, guild_name)
            hill_entry = hill_map.setdefault(hill_name, {"count": 0, "power": 0, "guilds": {}})
            hill_entry["count"] += 1
            hill_entry["power"] += member["power"]
            guild_entry = hill_entry["guilds"].setdefault(guild_key, {
                "name": guild_name,
                "code": guild_code,
                "prefix": guild_prefix,
                "display_name": build_guild_display_name(guild_code, guild_prefix, guild_name),
                "count": 0,
                "power": 0,
                "custom_power": 0,
                "top_member": None,
                "leader_name": "",
                "updated_at": "",
            })
            guild_entry["count"] += 1
            guild_entry["power"] += member["power"]
            guild_entry["custom_power"] = max(guild_entry["custom_power"], int(member.get("guild_power", 0) or 0))
            guild_entry["updated_at"] = max(guild_entry.get("updated_at", ""), member.get("updated_at", "") or member.get("created_at", ""))
            if member.get("role") == "盟主" and member.get("name"):
                guild_entry["leader_name"] = member["name"]
            if guild_entry["top_member"] is None or member["power"] > guild_entry["top_member"]["power"]:
                guild_entry["top_member"] = member

        hills = []
        flat_guilds = []
        for hill_name, hill_info in sorted(hill_map.items(), key=lambda item: item[1]["power"], reverse=True):
            guilds = []
            for _, guild_info in sorted(hill_info["guilds"].items(), key=lambda item: item[1]["power"], reverse=True):
                guild_payload = {
                    "key": build_guild_key(guild_info["code"], guild_info["prefix"], guild_info["name"]),
                    "name": guild_info["name"],
                    "code": guild_info["code"],
                    "prefix": guild_info["prefix"],
                    "display_name": guild_info["display_name"],
                    "hill": hill_name,
                    "count": guild_info["count"],
                    "power": guild_info["custom_power"] or guild_info["power"],
                    "custom_power": guild_info["custom_power"],
                    "top_member": serialize_member(guild_info["top_member"]) if guild_info["top_member"] else None,
                    "leader_name": guild_info.get("leader_name", ""),
                    "updated_at": guild_info.get("updated_at", ""),
                }
                guilds.append(guild_payload)
                flat_guilds.append(guild_payload)
            hills.append({"name": hill_name, "count": hill_info["count"], "power": hill_info["power"], "guilds": guilds})

        ranking = [serialize_member(member) for member in sorted(members, key=lambda item: item["power"], reverse=True)[:10]]
        public_announcements = [
            {
                "id": row["id"],
                "title": row["title"],
                "content": row["content"],
                "category": row["category"],
                "created_at": row["created_at"],
                "author": row["author"],
            }
            for row in announcements
        ]

        return {
            "alliance_name": members[0]["alliance"] if len({member['alliance'] for member in members}) == 1 else "联盟总览",
            "member_count": len(members),
            "total_power": total_power,
            "guild_count": len(flat_guilds),
            "hill_count": len(hills),
            "top_member": ranking[0] if ranking else None,
            "top_guild": flat_guilds[0] if flat_guilds else None,
            "hills": hills,
            "guilds": flat_guilds,
            "ranking": ranking,
            "announcements": [item for item in public_announcements if item["category"] == "公告"],
            "melon_posts": [item for item in public_announcements if item["category"] == "瓜棚"],
        }

    def list_members(self, query):
        search = (query.get("search", [""])[0]).strip().lower()
        hill = (query.get("hill", ["all"])[0]).strip()
        guild = (query.get("guild", ["all"])[0]).strip()
        sort = (query.get("sort", ["power-desc"])[0]).strip()
        order_sql = {
            "power-desc": "power DESC, id DESC",
            "power-asc": "power ASC, id ASC",
            "name-asc": "name COLLATE NOCASE ASC, id DESC",
        }.get(sort, "power DESC, id DESC")

        clauses = []
        values = []
        if hill and hill != "all":
            clauses.append("hill = ?")
            values.append(hill)
        if guild and guild != "all":
            clauses.append("guild = ?")
            values.append(guild)
        if search:
            clauses.append("(LOWER(alliance) LIKE ? OR LOWER(hill) LIKE ? OR LOWER(guild) LIKE ? OR LOWER(name) LIKE ? OR LOWER(role) LIKE ? OR LOWER(realm) LIKE ? OR LOWER(pet) LIKE ? OR LOWER(note) LIKE ?)")
            keyword = f"%{search}%"
            values.extend([keyword] * 8)

        where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        sql = f"SELECT * FROM members {where_sql} ORDER BY {order_sql}"
        with open_db() as connection:
            rows = connection.execute(sql, values).fetchall()
        return {"items": [serialize_member(dict(row)) for row in rows]}

    def list_announcements(self):
        with open_db() as connection:
            rows = connection.execute("SELECT * FROM announcements ORDER BY id DESC").fetchall()
        return {
            "items": [
                {
                    "id": row["id"],
                    "title": row["title"],
                    "content": row["content"],
                    "category": row["category"],
                    "created_at": row["created_at"],
                    "author": row["author"],
                }
                for row in rows
            ]
        }

    def login(self, payload):
        username = str(payload.get("username", "")).strip()
        password = str(payload.get("password", "")).strip()
        with open_db() as connection:
            row = connection.execute("SELECT * FROM admins WHERE username = ?", (username,)).fetchone()

        if not row:
            self.send_json({"error": "账号或密码错误"}, status=HTTPStatus.UNAUTHORIZED)
            return

        calculated = hash_password(password, row["salt"])
        if not hmac.compare_digest(calculated, row["password_hash"]):
            self.send_json({"error": "账号或密码错误"}, status=HTTPStatus.UNAUTHORIZED)
            return

        token = secrets.token_hex(24)
        sessions[token] = {"admin_id": row["id"], "username": row["username"], "display_name": row["display_name"], "created_at": datetime.now().timestamp()}

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Set-Cookie", f"{SESSION_COOKIE}={token}; HttpOnly; Path=/; SameSite=Lax")
        self.end_headers()
        self.wfile.write(json.dumps({"message": "登录成功", "admin": {"username": row["username"], "display_name": row["display_name"]}}, ensure_ascii=False).encode("utf-8"))

    def logout(self):
        token = self.read_session_token()
        if token:
            sessions.pop(token, None)

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Set-Cookie", f"{SESSION_COOKIE}=; HttpOnly; Path=/; Max-Age=0; SameSite=Lax")
        self.end_headers()
        self.wfile.write(json.dumps({"message": "已退出登录"}, ensure_ascii=False).encode("utf-8"))

    def create_member(self, payload):
        try:
            member = validate_member(payload)
        except ValueError as error:
            self.send_json({"error": str(error)}, status=HTTPStatus.BAD_REQUEST)
            return
        timestamp = now_text()
        member["created_at"] = timestamp
        member["updated_at"] = timestamp
        with open_db() as connection:
            if not guild_exists(connection, member["guild_code"], member["guild_prefix"], member["guild"]):
                self.send_json({"error": "该妖盟不存在，请先创建妖盟"}, status=HTTPStatus.BAD_REQUEST)
                return
            if member_name_exists(connection, member["guild_code"], member["guild_prefix"], member["guild"], member["name"]):
                self.send_json({"error": "该成员在当前妖盟里已存在"}, status=HTTPStatus.CONFLICT)
                return
            cursor = connection.execute(
                """
                INSERT INTO members (
                    alliance, hill, guild_code, guild_prefix, guild_power, guild, name, role, realm, power, hp, attack, defense, speed, bonus_damage, damage_reduction, pet, note, screenshot_path, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    member["alliance"], member["hill"], member["guild_code"], member["guild_prefix"], member["guild_power"], member["guild"], member["name"], member["role"], member["realm"], member["power"],
                    member["hp"], member["attack"], member["defense"], member["speed"], member["bonus_damage"], member["damage_reduction"], member["pet"], member["note"],
                    member["screenshot_path"],
                    timestamp, timestamp,
                ),
            )
            upsert_guild_registry(connection, member)
            connection.commit()

        member["id"] = cursor.lastrowid
        self.send_json({"message": "成员创建成功", "item": member}, status=HTTPStatus.CREATED)

    def create_guild(self, payload):
        try:
            guild = validate_guild(payload)
        except ValueError as error:
            self.send_json({"error": str(error)}, status=HTTPStatus.BAD_REQUEST)
            return
        timestamp = now_text()
        guild["updated_at"] = timestamp
        guild_key = build_guild_key(guild["guild_code"], guild["guild_prefix"], guild["guild"])
        with open_db() as connection:
            if guild_exists(connection, guild["guild_code"], guild["guild_prefix"], guild["guild"]):
                self.send_json({"error": f"妖盟编号 {guild['guild_code'] or guild['guild']} 已存在，不能重复创建"}, status=HTTPStatus.CONFLICT)
                return
            exists = connection.execute("SELECT 1 FROM guild_registry WHERE guild_key = ?", (guild_key,)).fetchone()
            if exists:
                self.send_json({"error": "该妖盟已存在"}, status=HTTPStatus.CONFLICT)
                return
            connection.execute(
                """
                INSERT INTO guild_registry (
                    guild_key, alliance, hill, guild_code, guild_prefix, guild, guild_power, leader_name, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    guild_key,
                    guild["alliance"],
                    guild["hill"],
                    guild["guild_code"],
                    guild["guild_prefix"],
                    guild["guild"],
                    guild["guild_power"],
                    guild["leader_name"],
                    timestamp,
                ),
            )
            connection.commit()
        self.send_json({"message": "妖盟创建成功", "item": {**guild, "guild_key": guild_key}}, status=HTTPStatus.CREATED)

    def update_member(self, member_id, payload):
        try:
            member = validate_member(payload)
        except ValueError as error:
            self.send_json({"error": str(error)}, status=HTTPStatus.BAD_REQUEST)
            return
        timestamp = now_text()
        with open_db() as connection:
            existing = connection.execute(
                "SELECT guild_code, guild_prefix, guild FROM members WHERE id = ?",
                (member_id,),
            ).fetchone()
            if not existing:
                self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            if not guild_exists(connection, member["guild_code"], member["guild_prefix"], member["guild"]):
                self.send_json({"error": "目标妖盟不存在，请先创建妖盟"}, status=HTTPStatus.BAD_REQUEST)
                return
            if member["name"] and member_name_exists(connection, member["guild_code"], member["guild_prefix"], member["guild"], member["name"], exclude_id=int(member_id)):
                self.send_json({"error": "该成员在当前妖盟里已存在"}, status=HTTPStatus.CONFLICT)
                return
            cursor = connection.execute(
                """
                UPDATE members
                SET alliance = ?, hill = ?, guild_code = ?, guild_prefix = ?, guild_power = ?, guild = ?, name = ?, role = ?, realm = ?, power = ?, hp = ?, attack = ?, defense = ?, speed = ?, bonus_damage = ?, damage_reduction = ?, pet = ?, note = ?, screenshot_path = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    member["alliance"], member["hill"], member["guild_code"], member["guild_prefix"], member["guild_power"], member["guild"], member["name"], member["role"], member["realm"], member["power"],
                    member["hp"], member["attack"], member["defense"], member["speed"], member["bonus_damage"], member["damage_reduction"], member["pet"], member["note"],
                    member["screenshot_path"],
                    timestamp, member_id,
                ),
            )
            member["updated_at"] = timestamp
            upsert_guild_registry(connection, member)
            if existing:
                old_key = build_guild_key(existing["guild_code"], existing["guild_prefix"], existing["guild"])
                new_key = build_guild_key(member["guild_code"], member["guild_prefix"], member["guild"])
                if old_key != new_key:
                    remaining = connection.execute(
                        "SELECT COUNT(*) AS count FROM members WHERE guild_code = ? AND guild_prefix = ? AND guild = ?",
                        (existing["guild_code"], existing["guild_prefix"], existing["guild"]),
                    ).fetchone()["count"]
                    if remaining == 0:
                        connection.execute("DELETE FROM guild_registry WHERE guild_key = ?", (old_key,))
            connection.commit()

        if cursor.rowcount == 0:
            self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
            return

        member["id"] = int(member_id)
        member["updated_at"] = timestamp
        self.send_json({"message": "成员更新成功", "item": member})

    def update_guild(self, guild_key, payload):
        try:
            guild = validate_guild(payload)
        except ValueError as error:
            self.send_json({"error": str(error)}, status=HTTPStatus.BAD_REQUEST)
            return
        next_key = build_guild_key(guild["guild_code"], guild["guild_prefix"], guild["guild"])
        timestamp = now_text()
        with open_db() as connection:
            existing = connection.execute("SELECT * FROM guild_registry WHERE guild_key = ?", (guild_key,)).fetchone()
            if not existing:
                self.send_json({"error": "妖盟不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            duplicate_code = guild_exists(connection, guild["guild_code"], guild["guild_prefix"], guild["guild"], exclude_key=guild_key)
            if duplicate_code:
                self.send_json({"error": f"妖盟编号 {guild['guild_code'] or guild['guild']} 已存在，不能重复保存"}, status=HTTPStatus.CONFLICT)
                return
            if next_key != guild_key:
                conflict = connection.execute("SELECT 1 FROM guild_registry WHERE guild_key = ?", (next_key,)).fetchone()
                if conflict:
                    self.send_json({"error": "新的妖盟标识已存在"}, status=HTTPStatus.CONFLICT)
                    return
                connection.execute("DELETE FROM guild_registry WHERE guild_key = ?", (guild_key,))
            connection.execute(
                """
                INSERT INTO guild_registry (
                    guild_key, alliance, hill, guild_code, guild_prefix, guild, guild_power, leader_name, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(guild_key) DO UPDATE SET
                    alliance = excluded.alliance,
                    hill = excluded.hill,
                    guild_code = excluded.guild_code,
                    guild_prefix = excluded.guild_prefix,
                    guild = excluded.guild,
                    guild_power = excluded.guild_power,
                    leader_name = excluded.leader_name,
                    updated_at = excluded.updated_at
                """,
                (
                    next_key,
                    guild["alliance"],
                    guild["hill"],
                    guild["guild_code"],
                    guild["guild_prefix"],
                    guild["guild"],
                    guild["guild_power"],
                    guild["leader_name"],
                    timestamp,
                ),
            )
            connection.execute(
                """
                UPDATE members
                SET alliance = ?, hill = ?, guild_code = ?, guild_prefix = ?, guild = ?, guild_power = ?, updated_at = ?
                WHERE guild_code = ? AND guild_prefix = ? AND guild = ?
                """,
                (
                    guild["alliance"],
                    guild["hill"],
                    guild["guild_code"],
                    guild["guild_prefix"],
                    guild["guild"],
                    guild["guild_power"],
                    timestamp,
                    existing["guild_code"],
                    existing["guild_prefix"],
                    existing["guild"],
                ),
            )
            if guild["leader_name"]:
                connection.execute(
                    """
                    UPDATE members
                    SET name = ?, updated_at = ?
                    WHERE guild_code = ? AND guild_prefix = ? AND guild = ? AND id = (
                        SELECT id FROM members
                        WHERE guild_code = ? AND guild_prefix = ? AND guild = ?
                        ORDER BY power DESC, id ASC LIMIT 1
                    )
                    """,
                    (
                        guild["leader_name"],
                        timestamp,
                        guild["guild_code"],
                        guild["guild_prefix"],
                        guild["guild"],
                        guild["guild_code"],
                        guild["guild_prefix"],
                        guild["guild"],
                    ),
                )
            connection.commit()
        self.send_json({"message": "妖盟更新成功", "item": {**guild, "guild_key": next_key}})

    def delete_guild(self, guild_key):
        with open_db() as connection:
            existing = connection.execute("SELECT * FROM guild_registry WHERE guild_key = ?", (guild_key,)).fetchone()
            if not existing:
                self.send_json({"error": "妖盟不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            connection.execute(
                "DELETE FROM members WHERE guild_code = ? AND guild_prefix = ? AND guild = ?",
                (existing["guild_code"], existing["guild_prefix"], existing["guild"]),
            )
            connection.execute("DELETE FROM guild_registry WHERE guild_key = ?", (guild_key,))
            connection.commit()
        self.send_json({"message": "妖盟删除成功"})

    def create_announcement(self, payload):
        title = str(payload.get("title", "")).strip()
        content = str(payload.get("content", "")).strip()
        category = str(payload.get("category", "公告")).strip() or "公告"
        if not title or not content:
            self.send_json({"error": "标题和内容不能为空"}, status=HTTPStatus.BAD_REQUEST)
            return
        if category not in {"公告", "瓜棚"}:
            self.send_json({"error": "分类不正确"}, status=HTTPStatus.BAD_REQUEST)
            return

        timestamp = now_text()
        with open_db() as connection:
            cursor = connection.execute(
                "INSERT INTO announcements (title, content, category, created_at, author) VALUES (?, ?, ?, ?, ?)",
                (title, content, category, timestamp, ""),
            )
            connection.commit()

        self.send_json(
            {"message": "内容发布成功", "item": {"id": cursor.lastrowid, "title": title, "content": content, "category": category, "created_at": timestamp}},
            status=HTTPStatus.CREATED,
        )
    def create_melon_post(self, user, payload):
        """Create a melon post - called from /api/melon endpoint."""
        title = str(payload.get("title", "")).strip()
        content = str(payload.get("content", "")).strip()
        if not title or not content:
            self.send_json({"error": "标题和内容不能为空"}, status=HTTPStatus.BAD_REQUEST)
            return

        # Get author name from user object
        author_name = user.get("username") or user.get("display_name") or "匿名用户"
        user_id = user.get("id") or user.get("admin_id")

        # Create melon post
        melon_item = create_melon_post(title, content, author_name, user_id)

        self.send_json(
            {"message": "瓜棚发布成功", "item": melon_item},
            status=HTTPStatus.CREATED,
        )

        try:
            broadcast_melon_post(melon_item)
        except Exception:
            pass

    def delete_melon_post(self, user, melon_id):
        """Delete a melon post - called from /api/melon/<id> endpoint."""
        if not melon_id.isdigit():
            self.send_json({"error": "参数无效"}, status=HTTPStatus.BAD_REQUEST)
            return

        user_id = user.get("id") or user.get("admin_id")
        username = user.get("username") or user.get("display_name") or ""

        with open_db() as connection:
            # Get the melon post
            row = connection.execute(
                "SELECT id, title, author, created_at FROM announcements WHERE id = ? AND category = '瓜棚'",
                (melon_id,),
            ).fetchone()

            if not row:
                self.send_json({"error": "瓜棚动态不存在"}, status=HTTPStatus.NOT_FOUND)
                return

            # Check if the user is the author (compare by username/display_name)
            author = row["author"] or ""
            if author != username:
                self.send_json({"error": "只有发布者才能撤回"}, status=HTTPStatus.FORBIDDEN)
                return

            # Check if within 2 minutes
            created_at = row["created_at"]
            try:
                created_time = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
                elapsed = (datetime.now() - created_time).total_seconds()
                if elapsed > 120:  # 2 minutes = 120 seconds
                    self.send_json({"error": "已超过2分钟，无法撤回"}, status=HTTPStatus.FORBIDDEN)
                    return
            except (ValueError, TypeError):
                self.send_json({"error": "时间解析错误"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # Delete the melon post
            cursor = connection.execute(
                "DELETE FROM announcements WHERE id = ?",
                (melon_id,),
            )
            connection.commit()

        if cursor.rowcount == 0:
            self.send_json({"error": "删除失败"}, status=HTTPStatus.NOT_FOUND)
            return

        try:
            broadcast_melon_deleted(int(melon_id))
        except Exception:
            pass

        self.send_json({"message": "撤回成功", "deleted_id": int(melon_id)})

    def update_announcement(self, announcement_id, payload):
        title = str(payload.get("title", "")).strip()
        content = str(payload.get("content", "")).strip()
        category = str(payload.get("category", "公告")).strip() or "公告"
        if not title or not content:
            self.send_json({"error": "标题和内容不能为空"}, status=HTTPStatus.BAD_REQUEST)
            return

        with open_db() as connection:
            cursor = connection.execute(
                "UPDATE announcements SET title = ?, content = ?, category = ? WHERE id = ?",
                (title, content, category, announcement_id),
            )
            connection.commit()

        if cursor.rowcount == 0:
            self.send_json({"error": "内容不存在"}, status=HTTPStatus.NOT_FOUND)
            return

        self.send_json({"message": "内容更新成功"})

    def get_member_screenshot(self, member_id):
        if not member_id.isdigit():
            return {"error": "参数无效"}
        with open_db() as connection:
            row = connection.execute("SELECT id, name, screenshot_path FROM members WHERE id = ?", (member_id,)).fetchone()
        if not row:
            return {"error": "成员不存在"}
        screenshot_url = row["screenshot_path"] or ""
        return {
            "member_id": row["id"],
            "member_name": row["name"],
            "screenshot_url": screenshot_url,
            "has_screenshot": bool(screenshot_url),
        }

    def upload_member_screenshot(self, member_id):
        if not member_id.isdigit():
            self.send_json({"error": "参数无效"}, status=HTTPStatus.BAD_REQUEST)
            return

        with open_db() as connection:
            row = connection.execute("SELECT id, screenshot_path FROM members WHERE id = ?", (member_id,)).fetchone()
        if not row:
            self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
            return

        form = FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            },
        )
        file_item = form["screenshot"] if "screenshot" in form else None
        if file_item is None or getattr(file_item, "file", None) is None:
            self.send_json({"error": "请先选择截图文件"}, status=HTTPStatus.BAD_REQUEST)
            return

        raw = file_item.file.read()
        if not raw:
            self.send_json({"error": "截图文件不能为空"}, status=HTTPStatus.BAD_REQUEST)
            return
        if len(raw) > 8 * 1024 * 1024:
            self.send_json({"error": "截图不能超过 8MB"}, status=HTTPStatus.BAD_REQUEST)
            return

        filename = Path(getattr(file_item, "filename", "") or "")
        suffix = filename.suffix.lower()
        if suffix not in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
            self.send_json({"error": "仅支持 png、jpg、jpeg、gif、webp 图片"}, status=HTTPStatus.BAD_REQUEST)
            return

        self.delete_member_screenshot_file(row["screenshot_path"])
        version_token = datetime.now().strftime("%Y%m%d%H%M%S")
        relative_path = f"/uploads/member-screenshots/member-{member_id}-{version_token}{suffix}"
        target_path = PUBLIC_DIR / relative_path.lstrip("/")
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(raw)

        timestamp = now_text()
        with open_db() as connection:
            connection.execute(
                "UPDATE members SET screenshot_path = ?, updated_at = ? WHERE id = ?",
                (relative_path, timestamp, member_id),
            )
            connection.commit()

        self.send_json(
            {
                "message": "成员截图上传成功",
                "item": {
                    "member_id": int(member_id),
                    "screenshot_url": relative_path,
                    "updated_at": timestamp,
                },
            }
        )

    def delete_member_screenshot(self, member_id):
        if not member_id.isdigit():
            self.send_json({"error": "参数无效"}, status=HTTPStatus.BAD_REQUEST)
            return

        with open_db() as connection:
            row = connection.execute("SELECT id, screenshot_path FROM members WHERE id = ?", (member_id,)).fetchone()
        if not row:
            self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
            return
        if not row["screenshot_path"]:
            self.send_json({"error": "该成员当前没有截图"}, status=HTTPStatus.BAD_REQUEST)
            return

        self.delete_member_screenshot_file(row["screenshot_path"])
        timestamp = now_text()
        with open_db() as connection:
            connection.execute(
                "UPDATE members SET screenshot_path = '', updated_at = ? WHERE id = ?",
                (timestamp, member_id),
            )
            connection.commit()

        self.send_json(
            {
                "message": "成员截图已删除",
                "item": {
                    "member_id": int(member_id),
                    "screenshot_url": "",
                    "updated_at": timestamp,
                },
            }
        )

    def import_members_from_excel(self):
        """从Excel文件导入成员数据"""
        form = FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            },
        )
        
        # 获取Excel文件
        file_item = form["file"] if "file" in form else None
        if file_item is None or getattr(file_item, "file", None) is None:
            self.send_json({"error": "请先选择Excel文件"}, status=HTTPStatus.BAD_REQUEST)
            return
        
        # 获取目标妖盟信息
        guild_code = str(form.getvalue("guild_code", "")).strip()
        guild_prefix = str(form.getvalue("guild_prefix", "")).strip()
        guild_name = str(form.getvalue("guild", "")).strip()
        
        if not guild_name:
            self.send_json({"error": "缺少妖盟信息"}, status=HTTPStatus.BAD_REQUEST)
            return
        
        # 读取文件内容
        raw = file_item.file.read()
        if not raw:
            self.send_json({"error": "Excel文件不能为空"}, status=HTTPStatus.BAD_REQUEST)
            return
        
        # 保存临时文件
        import tempfile
        import os
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name
        
        try:
            # 解析Excel
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active
            
            # 获取表头行（第一行）
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else "")
            
            # 找到各列的索引（列顺序不固定）
            col_indices = {}
            for i, header in enumerate(headers):
                header_lower = header.lower()
                if header_lower in ["名称", "昵称", "name"]:
                    col_indices["name"] = i
                elif header_lower in ["境界", "realm"]:
                    col_indices["realm"] = i
                elif header_lower in ["等级", "level", "lv", "role"]:
                    col_indices["role"] = i
                elif header_lower in ["敏捷", "speed"]:
                    col_indices["speed"] = i
                elif header_lower in ["战力", "power"]:
                    col_indices["power"] = i
                elif header_lower in ["灵兽", "宠物", "pet"]:
                    col_indices["pet"] = i
                elif header_lower in ["增伤", "bonus_damage", "bonus"]:
                    col_indices["bonus_damage"] = i
                elif header_lower in ["减伤", "damage_reduction", "reduction"]:
                    col_indices["damage_reduction"] = i
            
            if "name" not in col_indices:
                self.send_json({"error": "Excel中未找到成员名称列（名称/昵称）"}, status=HTTPStatus.BAD_REQUEST)
                return
            
            # 解析数据行
            rows_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = str(row[col_indices["name"]]).strip() if col_indices["name"] < len(row) and row[col_indices["name"]] else ""
                if not name:
                    continue
                
                row_data = {
                    "name": name,
                    "realm": str(row[col_indices["realm"]]).strip() if "realm" in col_indices and col_indices["realm"] < len(row) and row[col_indices["realm"]] else "待补充",
                    "role": str(row[col_indices["role"]]).strip() if "role" in col_indices and col_indices["role"] < len(row) and row[col_indices["role"]] else "成员",
                    "speed": row[col_indices["speed"]] if "speed" in col_indices and col_indices["speed"] < len(row) else 0,
                    "power": row[col_indices["power"]] if "power" in col_indices and col_indices["power"] < len(row) else 0,
                    "pet": str(row[col_indices["pet"]]).strip() if "pet" in col_indices and col_indices["pet"] < len(row) and row[col_indices["pet"]] else "待补充",
                    "bonus_damage": row[col_indices["bonus_damage"]] if "bonus_damage" in col_indices and col_indices["bonus_damage"] < len(row) else 0,
                    "damage_reduction": row[col_indices["damage_reduction"]] if "damage_reduction" in col_indices and col_indices["damage_reduction"] < len(row) else 0,
                }
                rows_data.append(row_data)
            
            if not rows_data:
                self.send_json({"error": "Excel中没有找到有效数据"}, status=HTTPStatus.BAD_REQUEST)
                return
            
            # Excel内部去重（按名称）
            seen_names = set()
            unique_rows = []
            for row_data in rows_data:
                if row_data["name"] not in seen_names:
                    seen_names.add(row_data["name"])
                    unique_rows.append(row_data)
            
            excel_duplicates = len(rows_data) - len(unique_rows)
            
            # 获取数据库中该妖盟已有的成员名称
            with open_db() as connection:
                existing_members = connection.execute(
                    "SELECT name FROM members WHERE guild_code = ? AND guild_prefix = ? AND guild = ?",
                    (guild_code, guild_prefix, guild_name)
                ).fetchall()
                existing_names = set(row["name"] for row in existing_members)
                
                # 检查妖盟是否存在
                guild = connection.execute(
                    "SELECT * FROM guild_registry WHERE guild_code = ? AND guild_prefix = ? AND guild = ?",
                    (guild_code, guild_prefix, guild_name)
                ).fetchone()
                
                if not guild:
                    self.send_json({"error": f"妖盟 '{guild_name}' 不存在，请先创建妖盟"}, status=HTTPStatus.BAD_REQUEST)
                    return
                
                alliance = guild["alliance"]
                hill = guild["hill"]
                current = get_current_auth(self)
                user = current.get("user") or {}
                if not current.get("authenticated") or not self.can_access_alliance(user, alliance, "manage_members"):
                    self.send_json({"error": "当前账号没有导入该联盟成员的权限"}, status=HTTPStatus.FORBIDDEN)
                    return
                
                # 过滤掉已存在的成员
                new_members = [row for row in unique_rows if row["name"] not in existing_names]
                db_duplicates = len(unique_rows) - len(new_members)
                
                if not new_members:
                    self.send_json({
                        "message": "所有成员均已存在，无需导入",
                        "imported": 0,
                        "skipped_excel_duplicates": excel_duplicates,
                        "skipped_existing": db_duplicates,
                        "total_rows": len(rows_data)
                    })
                    return
                
                # 插入新成员
                timestamp = now_text()
                imported_count = 0
                for row_data in new_members:
                    try:
                        power = parse_scaled_number(row_data["power"], "战力")
                        speed = parse_scaled_number(row_data["speed"], "敏捷")
                        bonus_damage = parse_scaled_number(row_data["bonus_damage"], "增伤")
                        damage_reduction = parse_scaled_number(row_data["damage_reduction"], "减伤")
                    except (TypeError, ValueError):
                        power = 0
                        speed = 0
                        bonus_damage = 0
                        damage_reduction = 0
                    
                    connection.execute(
                        """
                        INSERT INTO members (
                            alliance, hill, guild_code, guild_prefix, guild_power, guild, name, role, realm, power, hp, attack, defense, speed, bonus_damage, damage_reduction, pet, note, screenshot_path, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            alliance, hill, guild_code, guild_prefix, 0, guild_name,
                            row_data["name"], row_data["role"], row_data["realm"], power,
                            0, 0, 0, speed, bonus_damage, damage_reduction,
                            row_data["pet"], "", "", timestamp, timestamp
                        )
                    )
                    imported_count += 1
                
                connection.commit()
            
            self.send_json({
                "message": f"成功导入 {imported_count} 名成员",
                "imported": imported_count,
                "skipped_excel_duplicates": excel_duplicates,
                "skipped_existing": db_duplicates,
                "total_rows": len(rows_data)
            })
            
        except Exception as exc:
            self.send_json({"error": f"解析Excel文件失败: {str(exc)}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
        finally:
            # 删除临时文件
            try:
                os.unlink(tmp_path)
            except:
                pass

    def get_known_alliances(self, connection):
        rows = connection.execute(
            """
            SELECT DISTINCT alliance
            FROM members
            WHERE alliance IS NOT NULL AND alliance != ''
            ORDER BY alliance COLLATE NOCASE ASC
            """
        ).fetchall()
        return [row["alliance"] for row in rows]

    def sync_verified_member_flags(self, connection):
        connection.execute(
            """
            UPDATE members
            SET verified = 1
            WHERE id IN (
                SELECT DISTINCT member_id FROM users WHERE member_id IS NOT NULL AND role = ?
            )
            """,
            (ROLE_VERIFIEDUSER,),
        )
        connection.commit()

    def list_member_cert_requests(self, current, mine=False, member_id=None):
        user = current.get("user") or {}
        role = user.get("role") or ROLE_GUEST
        params = []
        clauses = ["r.status = 'pending'"]
        if mine:
            clauses.append("r.user_id = ?")
            params.append(user.get("id"))
        elif role == ROLE_ALLIANCEADMIN:
            clauses.append("r.alliance = ?")
            params.append(user.get("alliance", ""))
        elif role != ROLE_SUPERADMIN:
            clauses.append("r.user_id = ?")
            params.append(user.get("id"))
        if member_id:
            clauses.append("r.member_id = ?")
            params.append(int(member_id))

        where_sql = " AND ".join(clauses)
        with open_db() as connection:
            rows = connection.execute(
                f"""
                SELECT r.*, u.username, u.email, u.role AS user_role, u.alliance AS user_alliance,
                       m.name AS member_name, m.guild AS guild_name, m.verified AS member_verified
                FROM member_cert_requests r
                LEFT JOIN users u ON u.id = r.user_id
                LEFT JOIN members m ON m.id = r.member_id
                WHERE {where_sql}
                ORDER BY r.id DESC
                """,
                tuple(params),
            ).fetchall()
        return {
            "items": [
                {
                    "id": row["id"],
                    "user_id": row["user_id"],
                    "member_id": row["member_id"],
                    "alliance": row["alliance"],
                    "status": row["status"],
                    "created_at": row["created_at"],
                    "reviewed_at": row["reviewed_at"],
                    "reviewer_id": row["reviewer_id"],
                    "username": row["username"],
                    "email": row["email"],
                    "display_name": row["username"],
                    "member_name": row["member_name"],
                    "guild_name": row["guild_name"],
                    "member_verified": row["member_verified"],
                }
                for row in rows
            ]
        }

    def create_member_cert_request(self, current, payload):
        user = current.get("user") or {}
        member_id = str(payload.get("member_id", "")).strip()
        if not member_id.isdigit():
            self.send_json({"error": "参数无效"}, status=HTTPStatus.BAD_REQUEST)
            return
        with open_db() as connection:
            member = connection.execute("SELECT * FROM members WHERE id = ?", (member_id,)).fetchone()
            if not member:
                self.send_json({"error": "成员不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            if int(member["verified"] or 0) == 1:
                self.send_json({"error": "该成员已经认证"}, status=HTTPStatus.CONFLICT)
                return
            exists = connection.execute(
                """
                SELECT id FROM member_cert_requests
                WHERE user_id = ? AND member_id = ? AND status = 'pending'
                """,
                (user.get("id"), member_id),
            ).fetchone()
            if exists:
                self.send_json({"error": "你已经提交过该成员的认证申请"}, status=HTTPStatus.CONFLICT)
                return
            timestamp = now_text()
            connection.execute(
                """
                INSERT INTO member_cert_requests (user_id, member_id, alliance, status, created_at)
                VALUES (?, ?, ?, 'pending', ?)
                """,
                (user.get("id"), member_id, member["alliance"], timestamp),
            )
            connection.commit()
        self.send_json({"message": "认证申请已提交"}, status=HTTPStatus.CREATED)

    def review_member_cert_request(self, request_id, payload, current):
        action = str(payload.get("action", "")).strip().lower()
        if action not in {"approve", "reject"}:
            self.send_json({"error": "操作无效"}, status=HTTPStatus.BAD_REQUEST)
            return
        if not str(request_id).isdigit():
            self.send_json({"error": "参数无效"}, status=HTTPStatus.BAD_REQUEST)
            return
        user = current.get("user") or {}
        with open_db() as connection:
            row = connection.execute(
                """
                SELECT r.*, u.username, u.role AS user_role, u.alliance AS user_alliance, m.verified AS member_verified
                FROM member_cert_requests r
                LEFT JOIN users u ON u.id = r.user_id
                LEFT JOIN members m ON m.id = r.member_id
                WHERE r.id = ?
                """,
                (request_id,),
            ).fetchone()
            if not row:
                self.send_json({"error": "申请不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            if current.get("user", {}).get("role") == ROLE_ALLIANCEADMIN and row["alliance"] != user.get("alliance", ""):
                self.send_json({"error": "只能审核本联盟的认证申请"}, status=HTTPStatus.FORBIDDEN)
                return
            if action == "approve":
                if int(row["member_verified"] or 0) == 1:
                    self.send_json({"error": "该成员已经认证"}, status=HTTPStatus.CONFLICT)
                    return
                timestamp = now_text()
                connection.execute(
                    "UPDATE members SET verified = 1, updated_at = ? WHERE id = ?",
                    (timestamp, row["member_id"]),
                )
                connection.execute(
                    """
                    UPDATE users
                    SET member_id = ?, member = ?, role = ?, alliance = ?
                    WHERE id = ?
                    """,
                    (row["member_id"], row["member_id"], ROLE_VERIFIEDUSER, row["alliance"], row["user_id"]),
                )
                update_user_sessions_for_user(
                    row["user_id"],
                    member_id=row["member_id"],
                    member=row["member_id"],
                    role=ROLE_VERIFIEDUSER,
                    alliance=row["alliance"],
                )
                connection.execute(
                    """
                    UPDATE member_cert_requests
                    SET status = 'approved', reviewed_at = ?, reviewer_id = ?
                    WHERE id = ?
                    """,
                    (timestamp, user.get("id"), request_id),
                )
                connection.commit()
                self.send_json({"message": "认证申请已通过"})
                return
            timestamp = now_text()
            connection.execute(
                """
                UPDATE member_cert_requests
                SET status = 'rejected', reviewed_at = ?, reviewer_id = ?
                WHERE id = ?
                """,
                (timestamp, user.get("id"), request_id),
            )
            connection.commit()
        self.send_json({"message": "认证申请已拒绝"})

    def list_admin_role_requests(self, mark_read=False):
        with open_db() as connection:
            unread_count = connection.execute(
                "SELECT COUNT(*) AS count FROM admin_role_requests WHERE status = 'pending' AND is_read = 0"
            ).fetchone()["count"]
            if mark_read and unread_count:
                connection.execute(
                    """
                    UPDATE admin_role_requests
                    SET is_read = 1, read_at = ?
                    WHERE status = 'pending' AND is_read = 0
                    """,
                    (now_text(),),
                )
                connection.commit()
            rows = connection.execute(
                """
                SELECT r.*, u.username, u.email, u.role AS current_role
                FROM admin_role_requests r
                LEFT JOIN users u ON u.id = r.user_id
                WHERE r.status = 'pending'
                ORDER BY r.id DESC
                """
            ).fetchall()
        return {
            "items": [
                {
                    "id": row["id"],
                    "user_id": row["user_id"],
                    "username": row["username"],
                    "email": row["email"],
                    "current_role": row["current_role"],
                    "alliance": row["alliance"],
                    "status": row["status"],
                    "is_read": int(row["is_read"] or 0),
                    "created_at": row["created_at"],
                }
                for row in rows
            ],
            "unread_count": 0 if mark_read else unread_count,
        }

    def create_admin_role_request(self, current, payload):
        user = current.get("user") or {}
        alliance = str(payload.get("alliance", "")).strip()
        if not alliance:
            self.send_json({"error": "请选择联盟"}, status=HTTPStatus.BAD_REQUEST)
            return
        with open_db() as connection:
            known_alliances = set(self.get_known_alliances(connection))
            if alliance not in known_alliances:
                self.send_json({"error": "联盟不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            row = connection.execute("SELECT role, alliance FROM users WHERE id = ?", (user.get("id"),)).fetchone()
            if not row:
                self.send_json({"error": "用户不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            if row["role"] in {ROLE_ALLIANCEADMIN, ROLE_SUPERADMIN}:
                self.send_json({"error": "当前角色无需再次申请"}, status=HTTPStatus.CONFLICT)
                return
            exists = connection.execute(
                """
                SELECT id FROM admin_role_requests
                WHERE user_id = ? AND status = 'pending'
                """,
                (user.get("id"),),
            ).fetchone()
            if exists:
                self.send_json({"error": "你已经有待审核的申请"}, status=HTTPStatus.CONFLICT)
                return
            timestamp = now_text()
            connection.execute(
                """
                INSERT INTO admin_role_requests (user_id, alliance, status, is_read, created_at)
                VALUES (?, ?, 'pending', 0, ?)
                """,
                (user.get("id"), alliance, timestamp),
            )
            connection.commit()
        self.send_json({"message": "联盟管理员申请已提交"}, status=HTTPStatus.CREATED)

    def review_admin_role_request(self, request_id, payload, current):
        action = str(payload.get("action", "")).strip().lower()
        if action not in {"approve", "reject"}:
            self.send_json({"error": "操作无效"}, status=HTTPStatus.BAD_REQUEST)
            return
        if not str(request_id).isdigit():
            self.send_json({"error": "参数无效"}, status=HTTPStatus.BAD_REQUEST)
            return
        with open_db() as connection:
            row = connection.execute(
                """
                SELECT r.*, u.username, u.role AS current_role
                FROM admin_role_requests r
                LEFT JOIN users u ON u.id = r.user_id
                WHERE r.id = ?
                """,
                (request_id,),
            ).fetchone()
            if not row:
                self.send_json({"error": "申请不存在"}, status=HTTPStatus.NOT_FOUND)
                return
            if action == "approve":
                timestamp = now_text()
                connection.execute(
                    """
                    UPDATE users
                    SET role = ?, alliance = ?
                    WHERE id = ?
                    """,
                    (ROLE_ALLIANCEADMIN, row["alliance"], row["user_id"]),
                )
                update_user_sessions_for_user(
                    row["user_id"],
                    role=ROLE_ALLIANCEADMIN,
                    alliance=row["alliance"],
                    is_admin=False,
                )
                connection.execute(
                    """
                    UPDATE admin_role_requests
                    SET status = 'approved', reviewed_at = ?, reviewer_id = ?
                    WHERE id = ?
                    """,
                    (timestamp, current.get("user", {}).get("id"), request_id),
                )
                connection.commit()
                self.send_json({"message": "联盟管理员申请已通过"})
                return
            timestamp = now_text()
            connection.execute(
                """
                UPDATE admin_role_requests
                SET status = 'rejected', reviewed_at = ?, reviewer_id = ?
                WHERE id = ?
                """,
                (timestamp, current.get("user", {}).get("id"), request_id),
            )
            connection.commit()
        self.send_json({"message": "联盟管理员申请已拒绝"})

    def get_current_user_profile(self):
        current = get_current_auth(self)
        if not current.get("authenticated"):
            return {"authenticated": False, "account": None, "member": None, "linked": False, "is_admin": False}
        if current.get("is_admin"):
            return {
                "authenticated": True,
                "account": current["user"],
                "member": None,
                "linked": False,
                "is_admin": True,
            }

        member = self.fetch_member_for_user(current["user"]["id"])
        return {
            "authenticated": True,
            "account": current["user"],
            "member": member,
            "linked": bool(member),
            "is_admin": False,
            "auto_link_hint": "用户名与成员昵称完全一致时，系统会自动关联成员档案。",
        }

    def fetch_member_for_user(self, user_id):
        with open_db() as connection:
            row = connection.execute(
                """
                SELECT m.*
                FROM users u
                LEFT JOIN members m ON m.id = COALESCE(u.member_id, u.member)
                WHERE u.id = ?
                """,
                (user_id,),
            ).fetchone()
        if not row or row["id"] is None:
            return None
        return serialize_member(dict(row))

    def update_current_user_profile(self, user, payload):
        member = self.fetch_member_for_user(user["id"])
        if not member:
            self.send_json({"error": "当前账号还没有关联成员档案，请先联系管理员确认，或使用与成员昵称一致的账号登录"}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            name = str(payload.get("name", member["name"])).strip() or member["name"]
            realm = str(payload.get("realm", member["realm"])).strip() or member["realm"]
            power = parse_scaled_number(payload.get("power", member["power"]), "战力")
            speed = parse_scaled_number(payload.get("speed", member["speed"]), "敏捷")
            bonus_damage = parse_scaled_number(payload.get("bonus_damage", member["bonus_damage"]), "增伤")
            damage_reduction = parse_scaled_number(payload.get("damage_reduction", member["damage_reduction"]), "减伤")
            pet = str(payload.get("pet", member["pet"])).strip() or member["pet"]
            note = str(payload.get("note", member["note"])).strip()
        except (TypeError, ValueError):
            self.send_json({"error": "个人信息里的数值格式不正确"}, status=HTTPStatus.BAD_REQUEST)
            return

        timestamp = now_text()
        with open_db() as connection:
            connection.execute(
                """
                UPDATE members
                SET name = ?, realm = ?, power = ?, speed = ?, bonus_damage = ?, damage_reduction = ?, pet = ?, note = ?, updated_at = ?
                WHERE id = ?
                """,
                (name, realm, power, speed, bonus_damage, damage_reduction, pet, note, timestamp, member["id"]),
            )
            connection.commit()

        self.send_json({"message": "个人信息更新成功", "item": self.fetch_member_for_user(user["id"])})

    def upload_current_user_screenshot(self, user):
        member = self.fetch_member_for_user(user["id"])
        if not member:
            self.send_json({"error": "当前账号还没有关联成员档案，暂时无法上传截图"}, status=HTTPStatus.BAD_REQUEST)
            return
        self.upload_member_screenshot(str(member["id"]))

    def delete_current_user_screenshot(self, user):
        member = self.fetch_member_for_user(user["id"])
        if not member:
            self.send_json({"error": "当前账号还没有关联成员档案，暂时无法删除截图"}, status=HTTPStatus.BAD_REQUEST)
            return
        self.delete_member_screenshot(str(member["id"]))

    def delete_by_id(self, table_name, record_id):
        if not record_id.isdigit():
            self.send_json({"error": "参数无效"}, status=HTTPStatus.BAD_REQUEST)
            return
        if table_name == "members":
            with open_db() as connection:
                row = connection.execute("SELECT screenshot_path FROM members WHERE id = ?", (record_id,)).fetchone()
            if row and row["screenshot_path"]:
                self.delete_member_screenshot_file(row["screenshot_path"])
        with open_db() as connection:
            cursor = connection.execute(f"DELETE FROM {table_name} WHERE id = ?", (record_id,))
            connection.commit()
        if cursor.rowcount == 0:
            self.send_json({"error": "数据不存在"}, status=HTTPStatus.NOT_FOUND)
            return
        self.send_json({"message": "删除成功"})

    def read_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length) if length > 0 else b"{}"
        try:
            return json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            self.send_json({"error": "JSON 格式不正确"}, status=HTTPStatus.BAD_REQUEST)
            raise

    def require_admin(self):
        current = get_current_auth(self)
        if not current.get("authenticated") or not current.get("is_admin"):
            self.send_json({"error": "请先登录管理员账号"}, status=HTTPStatus.UNAUTHORIZED)
            return None
        return current["user"]

    def has_permission(self, user, permission):
        if not user:
            return False
        if user.get("role") == ROLE_SUPERADMIN or user.get("is_admin"):
            return True
        return permission in set(user.get("permissions") or [])

    def can_access_alliance(self, user, alliance_name, permission):
        if not self.has_permission(user, permission):
            return False
        if user.get("role") == ROLE_SUPERADMIN or user.get("is_admin"):
            return True
        return str(user.get("alliance") or "").strip() == str(alliance_name or "").strip()

    def require_permission(self, permission, alliance_name=None, allow_admin_account=True):
        current = get_current_auth(self)
        if not current.get("authenticated"):
            self.send_json({"error": "请先登录"}, status=HTTPStatus.UNAUTHORIZED)
            return None
        user = current.get("user") or {}
        if not allow_admin_account and current.get("is_admin"):
            self.send_json({"error": "当前账号不能执行该操作"}, status=HTTPStatus.FORBIDDEN)
            return None
        allowed = self.has_permission(user, permission) if alliance_name is None else self.can_access_alliance(user, alliance_name, permission)
        if not allowed:
            self.send_json({"error": "当前账号没有执行该操作的权限"}, status=HTTPStatus.FORBIDDEN)
            return None
        return user

    def require_normal_user(self):
        current = get_current_auth(self)
        if not current.get("authenticated"):
            self.send_json({"error": "请先登录普通用户账号"}, status=HTTPStatus.UNAUTHORIZED)
            return None
        if current.get("is_admin"):
            self.send_json({"error": "管理员请使用管理员功能页面"}, status=HTTPStatus.FORBIDDEN)
            return None
        return current["user"]

    def get_current_admin(self):
        current = get_current_auth(self)
        if not current.get("authenticated") or not current.get("is_admin"):
            return None
        return current["user"]

    def get_current_user_or_admin(self):
        """Get current user whether admin or normal user."""
        current = get_current_auth(self)
        if not current.get("authenticated"):
            return None
        return current["user"]

    def get_guild_registry_item(self, guild_key):
        with open_db() as connection:
            row = connection.execute("SELECT * FROM guild_registry WHERE guild_key = ?", (guild_key,)).fetchone()
        return dict(row) if row else None

    def get_member_item(self, member_id):
        if not str(member_id).isdigit():
            return None
        with open_db() as connection:
            row = connection.execute("SELECT * FROM members WHERE id = ?", (int(member_id),)).fetchone()
        return dict(row) if row else None

    def read_session_token(self):
        cookie_header = self.headers.get("Cookie")
        if not cookie_header:
            return None
        cookie = SimpleCookie()
        cookie.load(cookie_header)
        morsel = cookie.get(SESSION_COOKIE)
        return morsel.value if morsel else None

    def delete_member_screenshot_file(self, screenshot_path):
        if not screenshot_path:
            return
        target = PUBLIC_DIR / str(screenshot_path).lstrip("/")
        try:
            resolved = target.resolve()
            uploads_root = UPLOADS_DIR.resolve()
            if str(resolved).startswith(str(uploads_root)) and resolved.exists():
                resolved.unlink()
        except FileNotFoundError:
            return

    def send_json(self, payload, status=HTTPStatus.OK):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, fmt, *args):
        sys.stdout.write("%s - - [%s] %s\n" % (self.address_string(), self.log_date_time_string(), fmt % args))


def validate_member(payload):
    allow_empty_name = bool(payload.get("allow_empty_name"))
    member = {
        "alliance": str(payload.get("alliance", "")).strip() or "🔮联盟",
        "hill": str(payload.get("hill", "")).strip() or "默认山头",
        "guild_code": str(payload.get("guild_code", "")).strip(),
        "guild_prefix": str(payload.get("guild_prefix", "")).strip(),
        "guild_power": parse_scaled_number(payload.get("guild_power", 0), "妖盟总战力"),
        "guild": str(payload.get("guild", "")).strip(),
        "name": str(payload.get("name", "")).strip(),
        "role": str(payload.get("role", "")).strip() or "成员",
        "realm": str(payload.get("realm", "")).strip() or "待补充",
        "power": parse_scaled_number(payload.get("power", 0), "战力"),
        "hp": parse_scaled_number(payload.get("hp", 0), "生命"),
        "attack": parse_scaled_number(payload.get("attack", 0), "攻击"),
        "defense": parse_scaled_number(payload.get("defense", 0), "防御"),
        "speed": parse_scaled_number(payload.get("speed", 0), "敏捷"),
        "bonus_damage": parse_scaled_number(payload.get("bonus_damage", 0), "增伤"),
        "damage_reduction": parse_scaled_number(payload.get("damage_reduction", 0), "减伤"),
        "pet": str(payload.get("pet", "")).strip() or "待补充",
        "note": str(payload.get("note", "")).strip(),
        "screenshot_path": str(payload.get("screenshot_path", "")).strip(),
    }
    if not member["guild"]:
        raise ValueError("妖盟不能为空")
    if not member["name"] and member["role"] != "盟主" and not allow_empty_name:
        raise ValueError("成员昵称不能为空")
    return member


def validate_guild(payload):
    guild = {
        "alliance": str(payload.get("alliance", "")).strip() or "🔮联盟",
        "hill": str(payload.get("hill", "")).strip() or str(payload.get("alliance", "")).strip() or "🔮联盟",
        "guild_code": str(payload.get("guild_code", "")).strip(),
        "guild_prefix": str(payload.get("guild_prefix", "")).strip(),
        "guild_power": parse_scaled_number(payload.get("guild_power", 0), "妖盟总战力"),
        "guild": str(payload.get("guild", "")).strip(),
        "leader_name": str(payload.get("leader_name", "")).strip(),
    }
    if not guild["guild"]:
        raise ValueError("妖盟不能为空")
    return guild


def serialize_member(member):
    return {
        "id": member["id"],
        "alliance": member["alliance"],
        "hill": member.get("hill", "默认山头"),
        "guild_code": member.get("guild_code", ""),
        "guild_prefix": member.get("guild_prefix", ""),
        "guild_power": member.get("guild_power", 0),
        "guild": member["guild"],
        "guild_display": build_guild_display_name(member.get("guild_code", ""), member.get("guild_prefix", ""), member["guild"]),
        "name": member["name"],
        "role": member["role"],
        "realm": member["realm"],
        "power": member["power"],
        "hp": member["hp"],
        "attack": member["attack"],
        "defense": member["defense"],
        "speed": member["speed"],
        "bonus_damage": member.get("bonus_damage", 0),
        "damage_reduction": member.get("damage_reduction", 0),
        "pet": member["pet"],
        "note": member["note"],
        "verified": bool(member.get("verified", 0)),
        "screenshot_url": member.get("screenshot_path", ""),
        "has_screenshot": bool(member.get("screenshot_path")),
        "created_at": member.get("created_at"),
        "updated_at": member.get("updated_at"),
    }


# ==================== WebSocket Support for Melon Posts ====================


def register_ws_client(client):
    with ws_clients_lock:
        ws_clients.add(client)


def unregister_ws_client(client):
    with ws_clients_lock:
        ws_clients.discard(client)


def snapshot_ws_clients():
    with ws_clients_lock:
        return list(ws_clients)


def broadcast_melon_post(melon_item):
    """Broadcast a new melon post to all connected WebSocket clients."""
    _broadcast_ws_text({
        "type": "melon_new",
        "id": melon_item.get("id"),
        "title": melon_item.get("title", ""),
    })


def broadcast_melon_deleted(deleted_id):
    """Broadcast a deleted melon post to all connected WebSocket clients."""
    _broadcast_ws_text({"type": "melon_deleted", "deleted_id": deleted_id})


def _broadcast_ws_text(payload):
    message = json.dumps(payload, ensure_ascii=False)
    dead_clients = []
    for client in snapshot_ws_clients():
        try:
            client.send_text(message)
        except Exception:
            dead_clients.append(client)
    for client in dead_clients:
        unregister_ws_client(client)


class MelonWebSocketClient:
    def __init__(self, connection, client_address):
        self.connection = connection
        self.client_address = client_address
        self.closed = False
        self.send_lock = threading.Lock()

    def _recv_exact(self, size):
        buffer = bytearray()
        while len(buffer) < size:
            chunk = self.connection.recv(size - len(buffer))
            if not chunk:
                raise ConnectionError("WebSocket connection closed")
            buffer.extend(chunk)
        return bytes(buffer)

    def _build_frame(self, payload_bytes, opcode=0x1):
        first_byte = 0x80 | (opcode & 0x0F)
        length = len(payload_bytes)
        if length < 126:
            header = struct.pack("!BB", first_byte, length)
        elif length < 65536:
            header = struct.pack("!BBH", first_byte, 126, length)
        else:
            header = struct.pack("!BBQ", first_byte, 127, length)
        return header + payload_bytes

    def send_text(self, text):
        if self.closed:
            return
        frame = self._build_frame(text.encode("utf-8"), opcode=0x1)
        with self.send_lock:
            self.connection.sendall(frame)

    def send_pong(self, payload=b""):
        if self.closed:
            return
        frame = self._build_frame(payload, opcode=0xA)
        with self.send_lock:
            self.connection.sendall(frame)

    def close(self):
        if self.closed:
            return
        self.closed = True
        try:
            with self.send_lock:
                self.connection.sendall(self._build_frame(b"", opcode=0x8))
        except Exception:
            pass
        try:
            self.connection.close()
        except Exception:
            pass

    def read_loop(self):
        try:
            while not self.closed:
                header = self._recv_exact(2)
                first_byte, second_byte = header[0], header[1]
                opcode = first_byte & 0x0F
                masked = bool(second_byte & 0x80)
                payload_length = second_byte & 0x7F

                if payload_length == 126:
                    payload_length = struct.unpack("!H", self._recv_exact(2))[0]
                elif payload_length == 127:
                    payload_length = struct.unpack("!Q", self._recv_exact(8))[0]

                mask_key = self._recv_exact(4) if masked else b""
                payload = self._recv_exact(payload_length) if payload_length else b""
                if masked and payload:
                    payload = bytes(byte ^ mask_key[i % 4] for i, byte in enumerate(payload))

                if opcode == 0x8:
                    break
                if opcode == 0x9:
                    self.send_pong(payload)
        except Exception:
            pass
        finally:
            self.close()


def async_save_melon_to_db(title, content, author_name):
    """Save melon post to database asynchronously (runs in thread pool)."""
    def _save():
        timestamp = now_text()
        with open_db() as connection:
            cursor = connection.execute(
                "INSERT INTO announcements (title, content, category, created_at, author) VALUES (?, ?, ?, ?, ?)",
                (title, content, "\u74dc\u68da", timestamp, author_name),
            )
            connection.commit()
            return {
                "id": cursor.lastrowid,
                "title": title,
                "content": content,
                "category": "瓜棚",
                "created_at": timestamp,
                "author": author_name,
            }
    future = db_executor.submit(_save)
    # We don't wait here - the caller gets the future if needed
    # But for optimistic updates, we'll return a temporary item immediately
    return future


def create_melon_post(title, content, author_name, user_id=None):
    """Create a melon post immediately (synchronous for API response)."""
    timestamp = now_text()
    with open_db() as connection:
        cursor = connection.execute(
            "INSERT INTO announcements (title, content, category, created_at, author) VALUES (?, ?, ?, ?, ?)",
            (title, content, "\u74dc\u68da", timestamp, author_name),
        )
        connection.commit()
        melon_item = {
            "id": cursor.lastrowid,
            "title": title,
            "content": content,
            "category": "瓜棚",
            "created_at": timestamp,
            "author": author_name,
            "author_id": user_id,
        }
    return melon_item


# ==================== Main ====================

def main():
    initialize_database()
    initialize_auth_database()
    server = ThreadingHTTPServer((HOST, PORT), AllianceHandler)
    print(f"寻道大千联盟信息站已启动: http://{HOST}:{PORT}")
    print("默认管理员账号: admin")
    print("默认管理员密码: 123456")
    server.serve_forever()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n服务器已停止")
